import openpyxl
import re
import os
from pathlib import Path
import pandas as pd
import numpy as np
import datetime as dt

from coeqwalpackage.metrics import create_subset_unit


def get_xl_sheetnames(xlfn):
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    return (sheet_names)


def read_from_excel(xlfn, tabname, topleft, bottomright, hdr=True, dtypes=[]):
    # open excel file
    wb = openpyxl.load_workbook(xlfn, data_only=True)
    # get the 'Inputs' tab
    sheet_names = wb.get_sheet_names()
    inputs_sheet = wb.get_sheet_by_name(tabname)
    topsplit = re.split('(\d+)', topleft)
    botsplit = re.split('(\d+)', bottomright)

    if hdr:
        hdr_block = inputs_sheet[topleft:(botsplit[0] + topsplit[1])]
        hdr_values = [[str(v.value) for v in v1] for v1 in hdr_block][0]
        data_block = inputs_sheet[topsplit[0] + str(int(topsplit[1]) + 1):bottomright]
    else:
        data_block = inputs_sheet[topleft:bottomright]
        hdr_values = None
    data_list = []
    # loop over months
    for i, row in enumerate(data_block):
        # loop over columns
        tmp = []
        if row[0].value == 'null':
            pass
        else:
            if not dtypes:  # list of dtypes is empty
                for j, cell in enumerate(row[:]):
                    tmp.append(str(cell.value))
            elif len(dtypes) != len(row):
                print("Wrong number of dtypes provided - returning values as strings")
            else:
                for j, cell in enumerate(row[:]):
                    dtypi = dtypes[j]
                    if dtypi[0:2] == 'dt':
                        parsetxt = dtypi[2:]
                        if type(cell.value) is not dt.datetime:  #check if it's already parsed as a datetime
                            tmp.append(dt.datetime.strptime(cell.value, parsetxt))
                        else:
                            tmp.append((cell.value))
                    elif dtypi[0:5] == 'float':
                        tmp.append(float(cell.value))
                    else:
                        tmp.append(str(cell.value))
            data_list.append(tmp)
    return ([hdr_values, data_list])


def read_init_file(CtrlFile, CtrlTab):
    # Hard-coded file structure
    ScenarioDirInd = 'B2'
    ScenarioListFileInd = 'B3'
    ScenarioListTabInd = 'B4'
    IndexMinInd = 'C5'  # first scenario index
    IndexMaxInd = 'D5'  # last scenario index
    ScenariosNameMinInd = 'C6'  # first dss dir index
    ScenariosNameMaxInd = 'D6'  # last dss dir index
    ScenariosDirMinInd = 'C7'  # first dss dir index
    ScenariosDirMaxInd = 'D7'  # last dss dir index
    DVDssPathMinInd = 'C8'  # first dss path index
    DVDssPathMaxInd = 'D8'  # last dss path index
    SVDssPathMinInd = 'C9'  # first dss path index
    SVDssPathMaxInd = 'D9'  # last dss path index
    StartMinInd = 'C10'  # first start date index
    StartMaxInd = 'D10'  # last start date index
    EndMinInd = 'C11'  # First end date index
    EndMaxInd = 'D11'  # Last end date index
    GroupDataDirInd = 'B12'  # Group data extraction directory
    VarFileNameInd = 'B13'  # Variable list file name
    VarFileTabInd = 'B14'  # Variable list file tab
    VarMinInd = 'C15'  # top left of variable name block
    VarMaxInd = 'D15'  # bottom right of variable name block
    ExtractionDirInd = 'B16'
    DemDelDirInd = 'B17'
    ModelFilesDirInd = 'B18'
    ModelSubDirInd = 'B19'
    DemandsFileInd = 'B20'
    DemandsTabInd = 'B21'
    DemMinInd = 'C22'  # top left of demand name block
    DemMaxInd = 'D22'  # bottom right of demand name block

    InflowDirInd = 'B23'
    InflowFileInd = 'B24'
    InflowTabInd = 'B25'
    InflowMinInd = 'C26'  # top left of demand name block
    InflowMaxInd = 'D26'  # bottom right of demand name block

    # Control File Example
    #Item	Name or description	Upper Left Cell	Lower Right Cell
    #Scenarios Directory	../../CalSim3_Model_Runs/Scenarios		
    #Scenario Listings File	coeqwal_cs3_scenario_listing_v2.xlsx		
    #Scenario Listings Tab	scenario_list		
    #Scenario Indices	Scenario identifiers	A1	A11
    #Scenario Directory Indices	Scenario directory names	C1	C11
    #DSS Path Indices	Dss path names	G1	G11
    #Start Date Indices	Start dates	H1	H11
    #Start Date Indices	End dates	I1	I11
    #Group Extraction Directory	Group_Data_Extraction		
    #Variables Listing File	trend_report_variables_v3.xlsx		
    #Variables List Tab	Variables List		E177
    #Variables List Indices	Variables List Block	D8	E177
    #Data Extraction Dir	Data_Extraction		
    #Model Files Dir	Model_Files		
    #Model Files SubDir	ModelFiles		

    # Read directory structure and contol file name
    Hdr, ScenarioDir = read_from_excel(CtrlFile, CtrlTab, ScenarioDirInd, ScenarioDirInd,
                                       hdr=False)  # Scenarios directory (../../CalSim3_Model_Runs/Scenarios in the current structure)
    ScenarioDir = ScenarioDir[0][0]
    Hdr, ScenarioListFile = read_from_excel(CtrlFile, CtrlTab, ScenarioListFileInd, ScenarioListFileInd,
                                            hdr=False)  # DSS file names Excel workbook
    ScenarioListFile = ScenarioListFile[0][0]
    ScenarioListPath = os.path.join(ScenarioDir, ScenarioListFile)  # path to DSS file names Excel workbook

    # Read file names and ranges for DSS and vars from control file
    Hdr, ScenarioListTab = read_from_excel(CtrlFile, CtrlTab, ScenarioListTabInd, ScenarioListTabInd,
                                           hdr=False)  # DSS file names Excel workbook Tab
    ScenarioListTab = ScenarioListTab[0][0]
    Hdr, IndexMin = read_from_excel(CtrlFile, CtrlTab, IndexMinInd, IndexMinInd, hdr=False)  # Scenario Index Name UL
    IndexMin = IndexMin[0][0]
    Hdr, IndexMax = read_from_excel(CtrlFile, CtrlTab, IndexMaxInd, IndexMaxInd, hdr=False)  # Scenario Index Name LR
    IndexMax = IndexMax[0][0]
    Hdr, NameMin = read_from_excel(CtrlFile, CtrlTab, ScenariosNameMinInd, ScenariosNameMinInd,
                                   hdr=False)  # Scenario Name UL
    NameMin = NameMin[0][0]
    Hdr, NameMax = read_from_excel(CtrlFile, CtrlTab, ScenariosNameMaxInd, ScenariosNameMaxInd,
                                   hdr=False)  # Scenario Name UL
    NameMax = NameMax[0][0]
    Hdr, DirMin = read_from_excel(CtrlFile, CtrlTab, ScenariosDirMinInd, ScenariosDirMinInd,
                                  hdr=False)  # Scenario Dir Name UL
    DirMin = DirMin[0][0]
    Hdr, DirMax = read_from_excel(CtrlFile, CtrlTab, ScenariosDirMaxInd, ScenariosDirMaxInd,
                                  hdr=False)  # Scenario Dir Name UL
    DirMax = DirMax[0][0]
    Hdr, DVDssMin = read_from_excel(CtrlFile, CtrlTab, DVDssPathMinInd, DVDssPathMinInd, hdr=False)  # DSS Path Name UL
    DVDssMin = DVDssMin[0][0]
    Hdr, DVDssMax = read_from_excel(CtrlFile, CtrlTab, DVDssPathMaxInd, DVDssPathMaxInd, hdr=False)  # DSS Path Name LR
    DVDssMax = DVDssMax[0][0]
    Hdr, SVDssMin = read_from_excel(CtrlFile, CtrlTab, SVDssPathMinInd, SVDssPathMinInd, hdr=False)  # DSS Path Name UL
    SVDssMin = SVDssMin[0][0]
    Hdr, SVDssMax = read_from_excel(CtrlFile, CtrlTab, SVDssPathMaxInd, SVDssPathMaxInd, hdr=False)  # DSS Path Name LR
    SVDssMax = SVDssMax[0][0]
    Hdr, StartMin = read_from_excel(CtrlFile, CtrlTab, StartMinInd, StartMinInd, hdr=False)  # Start Date UL
    StartMin = StartMin[0][0]
    Hdr, StartMax = read_from_excel(CtrlFile, CtrlTab, StartMaxInd, StartMaxInd, hdr=False)  # Start Date LR
    StartMax = StartMax[0][0]
    Hdr, EndMin = read_from_excel(CtrlFile, CtrlTab, EndMinInd, EndMinInd, hdr=False)  # Start Date UL
    EndMin = EndMin[0][0]
    Hdr, EndMax = read_from_excel(CtrlFile, CtrlTab, EndMaxInd, EndMaxInd, hdr=False)  # Start Date LR
    EndMax = EndMax[0][0]
    Hdr, GroupDataDirName = read_from_excel(CtrlFile, CtrlTab, GroupDataDirInd, GroupDataDirInd,
                                            hdr=False)  # directory name for group data extraction (Group_Data_Extraction in current structure)
    GroupDataDirName = GroupDataDirName[0][0]
    GroupDataDirPath = os.path.join(ScenarioDir,
                                    GroupDataDirName)  # group data extraction directory (../../CalSim3_Model_Runs/Scenarios/Group_Data_Extraction in the current structure)
    #print(GroupDataDirPath)
    Hdr, VarListFileName = read_from_excel(CtrlFile, CtrlTab, VarFileNameInd, VarFileNameInd,
                                           hdr=False)  # directory name for variable listing (trend_report_variables_v3.xlsx in current structure)
    VarListFileName = VarListFileName[0][0]
    Hdr, VarListTab = read_from_excel(CtrlFile, CtrlTab, VarFileTabInd, VarFileTabInd,
                                      hdr=False)  # tab for variable listing (TrendReportVars_CS3 in current structure)
    VarListTab = VarListTab[0][0]
    Hdr, VarMin = read_from_excel(CtrlFile, CtrlTab, VarMinInd, VarMinInd, hdr=False)  # variable listing UL
    VarMin = VarMin[0][0]
    Hdr, VarMax = read_from_excel(CtrlFile, CtrlTab, VarMaxInd, VarMaxInd, hdr=False)  # variable listing LR
    VarMax = VarMax[0][0]
    Hdr, ExtractionDir = read_from_excel(CtrlFile, CtrlTab, ExtractionDirInd, ExtractionDirInd,
                                         hdr=False)  #  Var extraction Dir Name
    ExtractionDir = ExtractionDir[0][0]
    Hdr, DemandDeliveryDir = read_from_excel(CtrlFile, CtrlTab, DemDelDirInd, DemDelDirInd,
                                             hdr=False)  #  Var extraction Dir Name
    DemandDeliveryDir = DemandDeliveryDir[0][0]
    Hdr, ModelFilesDir = read_from_excel(CtrlFile, CtrlTab, ModelFilesDirInd, ModelFilesDirInd,
                                         hdr=False)  #  Var extraction Dir Name
    ModelFilesDir = ModelFilesDir[0][0]
    Hdr, ModelSubDir = read_from_excel(CtrlFile, CtrlTab, ModelSubDirInd, ModelSubDirInd,
                                       hdr=False)  #  Var extraction SubDir Name
    ModelSubDir = ModelSubDir[0][0]
    Hdr, DemandFileName = read_from_excel(CtrlFile, CtrlTab, DemandsFileInd, DemandsFileInd,
                                          hdr=False)  # directory name for variable listing (trend_report_variables_v3.xlsx in current structure)
    DemandFileName = DemandFileName[0][0]
    Hdr, DemandFileTab = read_from_excel(CtrlFile, CtrlTab, DemandsTabInd, DemandsTabInd,
                                         hdr=False)  # tab for variable listing (TrendReportVars_CS3 in current structure)
    DemandFileTab = DemandFileTab[0][0]
    Hdr, DemMin = read_from_excel(CtrlFile, CtrlTab, DemMinInd, DemMinInd, hdr=False)  # variable listing UL
    DemMin = DemMin[0][0]
    Hdr, DemMax = read_from_excel(CtrlFile, CtrlTab, DemMaxInd, DemMaxInd, hdr=False)  # variable listing LR
    DemMax = DemMax[0][0]

    Hdr, InflowDir = read_from_excel(CtrlFile, CtrlTab, InflowDirInd, InflowDirInd,
                                     hdr=False)  #  Var extraction Dir Name
    InflowDir = InflowDir[0][0]
    Hdr, InflowFileName = read_from_excel(CtrlFile, CtrlTab, InflowFileInd, InflowFileInd,
                                          hdr=False)  # directory name for variable listing (trend_report_variables_v3.xlsx in current structure)
    InflowFileName = InflowFileName[0][0]
    Hdr, InflowFileTab = read_from_excel(CtrlFile, CtrlTab, InflowTabInd, InflowTabInd,
                                         hdr=False)  # tab for variable listing (TrendReportVars_CS3 in current structure)
    InflowFileTab = InflowFileTab[0][0]
    Hdr, InflowMin = read_from_excel(CtrlFile, CtrlTab, InflowMinInd, InflowMinInd, hdr=False)  # variable listing UL
    InflowMin = InflowMin[0][0]
    Hdr, InflowMax = read_from_excel(CtrlFile, CtrlTab, InflowMaxInd, InflowMaxInd, hdr=False)  # variable listing LR
    InflowMax = InflowMax[0][0]

    # Construct file and directory names
    # File and directory names
    ScenarioListFileCsv = ScenarioListFile.replace(".xlsx", ".csv")
    DVDssNamesOut = 'DVDssNamesFrom_' + ScenarioListFileCsv  # output DSS names CSV
    SVDssNamesOut = 'SVDssNamesFrom_' + ScenarioListFileCsv  # output DSS names CSV
    ScenarioIndicesOut = 'IndicesFrom_' + ScenarioListFileCsv  # output DSS indices CSV
    DssDirsOut = 'DirNamesFrom_' + ScenarioListFileCsv  # output directory names CSV
    DVDssNamesOutPath = os.path.join(GroupDataDirPath, DVDssNamesOut)  # output DSS names CSV path
    SVDssNamesOutPath = os.path.join(GroupDataDirPath, SVDssNamesOut)  # output DSS names CSV path
    ScenarioIndicesOutPath = os.path.join(GroupDataDirPath, ScenarioIndicesOut)  # output DSS index names CSV path
    DssDirsOutPath = os.path.join(GroupDataDirPath, DssDirsOut)  # output DSS dir names CSV path

    # list of relevant variables file, tab, and range (B & C parts)
    VarListName = os.path.splitext(VarListFileName)[0]  # variable names file without extension
    VarListExt = os.path.splitext(VarListFileName)[1]  # variable names file extension
    VarListFile = VarListName + VarListExt  # full file name
    VarListFileCsv = VarListFile.replace(".xlsx", ".csv")
    VarListPath = os.path.join(ScenarioDir, VarListFile)
    DemandFilePath = os.path.join(ScenarioDir, DemandFileName)
    VarOut = 'VarsFrom_' + VarListFileCsv  # output compund variable names CSV
    VarOutPath = os.path.join(GroupDataDirPath, VarOut)
    DataOut = 'DataFrom_' + VarListFileCsv  # file name for multi-study output CSV
    DataOutPath = os.path.join(GroupDataDirPath, DataOut)  # file name for multi-study output CSV path
    ConvertDataOut = 'ConvertDataFrom_' + VarListFileCsv  # file name for multi-study output CSV
    ConvertDataOutPath = os.path.join(GroupDataDirPath, ConvertDataOut)  # file name for multi-study output CSV path
    ExtractionSubDir = 'Variables_From_' + VarListName + '_' + VarListTab
    ExtractionSubPath = os.path.join(ExtractionDir, ExtractionSubDir)
    DemandDeliverySubPath = os.path.join(ExtractionDir, DemandDeliveryDir)
    ModelSubPath = os.path.join('Model_Files', 'DSS', 'output')
    InflowOutSubPath = os.path.join(ExtractionDir, InflowDir)
    InflowFilePath = os.path.join(ScenarioDir, InflowFileName)

    # debug print
    # print(ScenarioListFile)    
    # print(ScenarioListTab)    
    # print(ScenariosistPath)
    # print(DssNamesOutPath)
    # print(ScenarioIndicesOutPath)
    # print(DssDirsOutPath)
    # print(VarListPath)
    # print(VarOutPath)
    # print(DataOutPath)
    # print(ExtractionSubPath)
    # print(ModelSubPath)
    # print(GroupDataDirPath)

    # return info
    return ScenarioListFile, ScenarioListTab, ScenarioListPath, DVDssNamesOutPath, SVDssNamesOutPath, ScenarioIndicesOutPath, DssDirsOutPath, VarListPath, VarListFile, VarListTab, VarOutPath, DataOutPath, ConvertDataOutPath, ExtractionSubPath, DemandDeliverySubPath, ModelSubPath, GroupDataDirPath, ScenarioDir, DVDssMin, DVDssMax, SVDssMin, SVDssMax, NameMin, NameMax, DirMin, DirMax, IndexMin, IndexMax, StartMin, StartMax, EndMin, EndMax, VarMin, VarMax, DemandFilePath, DemandFileName, DemandFileTab, DemMin, DemMax, InflowOutSubPath, InflowFilePath, InflowFileName, InflowFileTab, InflowMin, InflowMax


def scenario_id(sid: int) -> str:
    return f"s{int(sid):04d}"


def build_scenario_labels_from_listing(
        list_path: str,
        list_file: str,
        list_tab: str,
        code_col: str = "Index",
        label_columns: tuple[str, ...] = ("StudyName", "ShortDescription",
                                          "GoogleDriveFolderName", "Index")
) -> dict[int, str]:
    # Resolve a usable Excel path with graceful fallbacks
    candidates = []
    if list_path and list_file:
        candidates.append(os.path.join(list_path, list_file))  # dir + file (normal case)
    if list_path:
        candidates.append(list_path)                           # list_path might already be the file
    if list_file:
        candidates.append(list_file)                           # list_file might be a full/relative path

    fpath = next((p for p in candidates if p and os.path.isfile(p)), None)
    if fpath is None:
        tried = ", ".join(repr(p) for p in candidates if p)
        raise FileNotFoundError(f"Could not locate Excel file. Tried: {tried}")

    df = pd.read_excel(fpath, sheet_name=list_tab)

    if code_col not in df.columns:
        raise KeyError(f"Expected column '{code_col}' in {fpath}::{list_tab}")

    sid = df[code_col].astype(str).str.extract(r"s(\d{4})", expand=False)
    df["scenario_id"] = pd.to_numeric(sid, errors="coerce")
    df = df.dropna(subset=["scenario_id"]).copy()
    df["scenario_id"] = df["scenario_id"].astype(int)

    cols = [c for c in label_columns if c in df.columns]
    if not cols:
        df["display_label"] = df[code_col].astype(str)
    else:
        display = df[cols[0]].astype(str)
        for c in cols[1:]:
            display = display.where(display.str.strip().ne("") & display.notna(), df[c].astype(str))
        df["display_label"] = display

    return dict(zip(df["scenario_id"], df["display_label"]))


def _safe_name(s: str) -> str:
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", str(s)).strip("_")


def plots_path_for_pair(group_dir: str, s1: int, s2: int, plot_type: str, scenario_names: dict[int, str] | None = None,
                        make_dirs: bool = True) -> str:

    def scenario_id(s: int) -> str:
        return f"s{int(s):04d}"

    base = os.path.join(group_dir, "plots_output")
    pair_seg = f"{scenario_id(s1)}_vs_{scenario_id(s2)}"
    out = os.path.join(base, pair_seg, plot_type)

    if make_dirs:
        os.makedirs(out, exist_ok=True)
    return out


def plots_path_for_multi(base_dir, scenario_ids, plot_type, make_dirs: bool = True) -> str:
    """
    Build an output directory for multi-scenario plots under base_dir/<plot_type>/<sXXXX_...>.
    """
    base = Path(base_dir)
    scenario_seg = "_".join(scenario_id(int(s)) for s in scenario_ids)
    out = base / plot_type / scenario_seg
    if make_dirs:
        out.mkdir(parents=True, exist_ok=True)
    return str(out)


def tucp_years_for(sid: int, policy_map: dict, year_sets: dict[str, list[int] | None]) -> list[int] | None:
    mode = policy_map.get(int(sid), policy_map.get("default"))
    if mode is None:
        raise KeyError("tucp_years_for: policy_map must include 'default' or a per-scenario mapping.")
    if mode == "none":
        return None
    if mode not in year_sets:
        raise ValueError(f"Unknown TUCP mode '{mode}'. Valid keys in year_sets={list(year_sets.keys())}")
    return year_sets[mode]

def selected_tucp_years(
        df: pd.DataFrame,
        *,
        scenario: int,
        tucp_var_base: str = "TUCP_TRIGGER_DV",
        tucp_wy_month_count: int = 1
) -> list[int]:
    """
    Determine which water years are 'TUCP years' for a given scenario.

    Rules:
      - We look for the trigger column whose Part B ends with '..._sXXXX'
        and contains tucp_var_base (default: 'TUCP_TRIGGER_DV').
      - A month is considered 'triggered' if the value >= 1 (numeric).
      - A water year (Oct–Sep) is considered TUCP if it has >= tucp_wy_month_count
        triggered months.

    Returns: sorted list of water years (ints).
    """
    if not isinstance(df.index, pd.DatetimeIndex):
        raise TypeError("selected_tucp_years: df.index must be a DatetimeIndex")

    suffix = f"s{int(scenario):04d}"
    lvl1 = df.columns.get_level_values(1).astype(str)
    mask = lvl1.str.contains(tucp_var_base, regex=False) & lvl1.str.endswith(suffix)

    if mask.sum() == 0:
        raise KeyError(
            f"selected_tucp_years: could not find trigger column for scenario={scenario} "
            f"with base='{tucp_var_base}'. Expected something like '{tucp_var_base}_{suffix}' in level-1."
        )
    if mask.sum() > 1:
        # If there are multiple candidates, pick the first; force explicitness if needed later
        # (keeps this helper simple/robust).
        pass

    trig = pd.to_numeric(df.loc[:, mask].iloc[:, 0], errors="coerce")
    triggered = (trig >= 1)

    wy = df.index.year + (df.index.month >= 10)
    months_per_wy = triggered.groupby(wy).sum()
    wys = months_per_wy.index[months_per_wy >= int(tucp_wy_month_count)]
    return sorted([int(y) for y in wys])


def create_subset_tucp(
        df: pd.DataFrame,
        *,
        scenario: int,
        tucp_var_base: str = "TUCP_TRIGGER_DV",
        tucp_wy_month_count: int = 1
) -> pd.DataFrame:
    """
    Return a copy of df that contains only rows whose WaterYear is classified as TUCP
    for the given scenario (per 'selected_tucp_years' rules). All columns are retained.
    """
    wys = selected_tucp_years(
        df, scenario=scenario,
        tucp_var_base=tucp_var_base,
        tucp_wy_month_count=tucp_wy_month_count
    )
    if len(wys) == 0:
        # Hard-fail per your preference to avoid silent plots
        raise ValueError(
            f"create_subset_tucp: no TUCP water years found for scenario={scenario} "
            f"(base={tucp_var_base}, wy_month_count={tucp_wy_month_count})."
        )

    wy_index = df.index.year + (df.index.month >= 10)
    keep_mask = pd.Series(wy_index).isin(wys).to_numpy()
    return df.loc[keep_mask].copy()


def per_scenario_series(
        df: pd.DataFrame,
        *,
        varname: str,
        units: str,
        scenarios: list[int],
        use_tucp: bool = False,
        tucp_var_base: str = "TUCP_TRIGGER_DV",
        tucp_wy_month_count: int = 1,
        use_wyt: bool = False,
        wyt: list[int] | None = None,
        wyt_month: int | None = None,
        months: list[int] | None = None
) -> dict[int, pd.Series]:
    """
    Return one aggregated Series per scenario (sum across matching subcomponents) after
    applying optional TUCP/WYT/month filters.

    - TUCP filtering is *per scenario* (each scenario uses its own trigger var).
    - WYT filtering uses your existing create_subset_unit(..., water_year_type=wyt, month=wyt_month).
    - 'months' can further filter the final monthly series.

    Returns: {scenario_id -> pd.Series}, aligned to whatever index survives the filters.
    """
    out: dict[int, pd.Series] = {}

    for sid in scenarios:
        work_df = df
        if use_tucp:
            work_df = create_subset_tucp(
                df,
                scenario=sid,
                tucp_var_base=tucp_var_base,
                tucp_wy_month_count=tucp_wy_month_count
            )

        # Apply variable/unit + (optional) WYT filter
        if use_wyt:
            if wyt is None or wyt_month is None:
                raise ValueError("per_scenario_series: use_wyt=True requires both wyt and wyt_month.")
            var_df = create_subset_unit(work_df, varname, units, water_year_type=wyt, month=wyt_month)
        else:
            var_df = create_subset_unit(work_df, varname, units)

        if var_df.empty:
            raise KeyError(
                f"per_scenario_series: no data after var/unit selection "
                f"(var='{varname}', units='{units}')."
            )

        suffix = f"s{int(sid):04d}"
        if hasattr(var_df.columns, "levels") and len(getattr(var_df.columns, "levels", [])) >= 2:
            keep_cols = [c for c in var_df.columns if str(c[1]).endswith(suffix)]
        else:
            keep_cols = [c for c in var_df.columns if str(c).endswith(suffix)]

        if len(keep_cols) == 0:
            raise KeyError(
                f"per_scenario_series: no columns found for scenario={sid} "
                f"(var='{varname}', units='{units}')."
            )

        series = var_df[keep_cols].sum(axis=1)  # Sum subcomponents if there are multiple
        if months is not None:
            series = series[series.index.month.isin(months)]

        series.name = f"{varname}_{units}_s{sid:04d}"
        out[sid] = series

    return out


def years_triggered_by(
        df: pd.DataFrame,
        scenario: int,
        tucp_var: str = "TUCP_TRIGGER_DV",
        tucp_wy_month_count: int = 1,
        trigger_threshold: float = 1.0,
) -> list[int]:
    """
    Return a sorted list of Water Years where the TUCP trigger for the given scenario
    is active in >= tucp_wy_month_count months.

    Assumptions:
    - df has a 7-level MultiIndex on columns (CALSIM/..., Part B at level 1).
    - The TUCP trigger column contains `tucp_var` in Part B and ends with `s{scenario:04d}`.
    - A month is considered triggered if value >= trigger_threshold (default 1.0).
    - Water Year = Oct (prev year) ... Sep (this year).
    """
    # Find the trigger column for this scenario
    suffix = f"s{scenario:04d}"
    lvl1 = df.columns.get_level_values(1).astype(str)
    mask = lvl1.str.contains(str(tucp_var), regex=False) & lvl1.str.endswith(suffix)

    if not mask.any():
        # Let it error loudly if not found (your requested behavior)
        raise AttributeError(
            f"No TUCP trigger column found for scenario {scenario} with base '{tucp_var}'."
        )

    # If multiple columns match, use the first
    trigger_col = df.columns[mask][0]
    trigger_series = pd.to_numeric(df.loc[:, trigger_col], errors="coerce")

    # Boolean "triggered this month?"
    monthly_flag = trigger_series >= float(trigger_threshold)

    # Compute Water Year for the index (Oct–Sep)
    wy = df.index.year + (df.index.month >= 10)

    # Count triggered months per WY
    months_per_wy = monthly_flag.groupby(wy).sum(min_count=1)
    sel_years = sorted(months_per_wy.index[months_per_wy >= int(tucp_wy_month_count)].tolist())

    return sel_years


def mask_by_years_per_scenario(
        var_df: pd.DataFrame,
        years_by_scenario: dict[int, list[int]],
) -> pd.DataFrame:
    """
    Given a variable-subset DataFrame (MultiIndex columns) that includes all scenarios,
    set values to NaN for rows (months) that are NOT in the scenario's allowed water years.

    This preserves the original index/columns, but different scenarios can keep different
    year sets (non-aligned years are allowed).
    """
    out = var_df.copy()
    # Compute Water Year on the index
    wy = var_df.index.year + (var_df.index.month >= 10)

    # For each scenario, restrict to its allowed WYs
    for s, allowed_wys in years_by_scenario.items():
        if not allowed_wys:  # None or empty -> keep all years
            continue
        suffix = f"s{s:04d}"
        scenario_cols = [c for c in out.columns if str(c[1]).endswith(suffix)]
        if not scenario_cols:
            # No columns for this scenario; keep moving
            continue

        keep_mask = wy.isin(allowed_wys)
        # Rows not in allowed WYs -> NaN for that scenario's columns
        out.loc[~keep_mask, scenario_cols] = np.nan

    return out


def plots_path_for_group(
        base_dir: str,
        scenario_list: list[int],
        plot_type: str,
) -> str:
    """
    Create and return a group plots path like:
      <base_dir>/plots_output/group_s0020_s0021_s0022/<plot_type>

    Always creates the directories if missing.
    """
    def _scenario_code(x: int) -> str:
        return f"s{x:04d}"

    plots_root = os.path.join(base_dir, "plots_output")
    group_name = "group_" + "_".join(_scenario_code(s) for s in scenario_list)
    out_dir = os.path.join(plots_root, group_name, plot_type)
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def cols_present(df, wanted: list[str], warn: bool = True) -> list[str]:
    """
    Filter a list of wanted column names to only those present in df.

    Parameters
    ----------
    df : pd.DataFrame
        DataFrame to check columns against.
    wanted : list[str]
        List of desired column names.
    warn : bool
        If True, print a warning about missing columns.

    Returns
    -------
    list[str]
        Subset of wanted that exists in df.columns.
    """
    present = [c for c in wanted if c in df.columns]
    missing = [c for c in wanted if c not in df.columns]
    if missing and warn:
        preview = missing[:5]
        suffix = ' ...' if len(missing) > 5 else ''
        print(f"[WARN] {len(missing)} missing columns: {preview}{suffix}")
    return present






